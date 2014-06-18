from openpyxl import load_workbook, cell
from openpyxl.cell import coordinate_from_string, column_index_from_string
import requests
from datetime import datetime, date, time

#open workbook and get sheet
wb = load_workbook('biorepositories-csv.xlsx')
ws = wb.get_active_sheet()

START_ROW = 2
END_ROW = ws.get_highest_row()
TITLE_COL = 'A'
ORG_COL = 'G'
LAST_COL = 'I'

def get_ord_val(s):
    position = []
    chars = []
    count = 0
    for c in s:
        g = ord(c)
        if g > 127:
            position.append(count)
            chars.append(g)
        count += 1
    print s
    print position
    print chars

# Find non-ascii unicode characters
for row in ws.range('%s%s:%s%s'%(TITLE_COL, START_ROW, LAST_COL, END_ROW)):
    for cell in row:
        if(cell.value):
            try:
                (cell.value).decode('ascii')
            except UnicodeEncodeError:
                get_ord_val(cell.value)

#for row in ws.range('%s%s:%s%s'%(TITLE_COL, START_ROW, TITLE_COL, END_ROW)):
#    for cell in row:
#        coordinate = coordinate_from_string(cell.get_coordinate())
#        org = ws['%s%s'%(ORG_COL, coordinate[1])].value
#        try:
#            if cell.value:
#                if org not in cell.value:
#                    partTitle = cell.value
#                    cell.value = org + ' - ' + partTitle
#            else:
#                cell.value = ws['%s%s'%(ORG_COL, coordinate[1])].value
#        except UnicodeEncodeError:
#            print "{}: Contains non-ascii char".format(cell.get_coordinate())
            
#wb.save('biorepositories-csv.xlsx')